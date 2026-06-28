"""
Partenon Scribe - Excel Templates
Generate reusable Excel templates for budgets, vendors and cash flow.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

# Optional dependency
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Partenon palette for Excel (dark blue, amber, cyan accents)
COLOR_DARK = "0B1F3A"
COLOR_ACCENT = "00D4FF"
COLOR_AMBER = "FFB800"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT = "E8F4F8"


def _style_header(cell):
    """Apply header style to a cell."""
    cell.font = Font(bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill(start_color=COLOR_DARK, end_color=COLOR_DARK, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _style_body(cell, align: str = "left"):
    """Apply body style to a cell."""
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


def _format_currency(cell):
    """Apply currency format."""
    cell.number_format = '$#,##0.00'


def _format_percent(cell):
    """Apply percentage format."""
    cell.number_format = '0.00%'


class Templates:
    """Generate finance Excel templates."""

    def __init__(self):
        self.errors: List[str] = []

    def _check_openpyxl(self) -> bool:
        if not OPENPYXL_AVAILABLE:
            self.errors.append("openpyxl is not installed")
            return False
        return True

    def create_budget(self, filepath: str = "budget.xlsx",
                      period: Optional[str] = None,
                      line_items: Optional[List[Dict[str, Any]]] = None) -> Optional[str]:
        """Create a budget vs actual Excel template."""
        if not self._check_openpyxl():
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"

        period = period or datetime.now().strftime("%B %Y")

        # Title
        ws["A1"] = "BUDGET VS ACTUAL"
        ws["A1"].font = Font(bold=True, size=18, color=COLOR_DARK)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws["A2"] = "Period:"
        ws["B2"] = period
        ws["A2"].font = Font(bold=True)

        # Headers
        headers = ["Line Item", "Budget", "Actual", "Difference", "% Variance", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            _style_header(cell)

        default_items = [
            {"line_item": "Materials", "budget": 30000},
            {"line_item": "Advertising", "budget": 12000},
            {"line_item": "Salaries", "budget": 45000},
            {"line_item": "Rent", "budget": 15000},
            {"line_item": "Services", "budget": 5000},
            {"line_item": "Freight", "budget": 5000},
        ]
        line_items = line_items or default_items

        start_row = 5
        for idx, item in enumerate(line_items):
            row = start_row + idx
            ws.cell(row=row, column=1, value=item.get("line_item", ""))
            ws.cell(row=row, column=2, value=item.get("budget", 0))
            ws.cell(row=row, column=3, value=item.get("actual", 0))
            ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
            ws.cell(row=row, column=5, value=f"=IF(B{row}=0,0,D{row}/B{row})")
            ws.cell(row=row, column=6, value=item.get("notes", ""))

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                _style_body(cell, align="left" if col in (1, 6) else "center")
                if col in (2, 3, 4):
                    _format_currency(cell)
                elif col == 5:
                    _format_percent(cell)

        # Totals
        total_row = start_row + len(line_items)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")

        for col in (2, 3, 4):
            cell = ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{total_row - 1})"
            )
            cell.font = Font(bold=True)
            _format_currency(cell)
            _style_body(cell, align="center")

        ws.cell(row=total_row, column=5, value=f"=IF(B{total_row}=0,0,D{total_row}/B{total_row})")
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        _format_percent(ws.cell(row=total_row, column=5))
        _style_body(ws.cell(row=total_row, column=5), align="center")

        # Widths
        widths = [28, 18, 14, 14, 15, 40]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[4].height = 25
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:F{total_row}"

        # Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Budget vs Actual"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Line Item"

        data = Reference(ws, min_col=2, min_row=4, max_row=total_row - 1, max_col=3)
        cats = Reference(ws, min_col=1, min_row=5, max_row=total_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4

        ws.add_chart(chart, "H4")

        wb.save(filepath)
        return filepath

    def create_vendors(self, filepath: str = "vendors.xlsx") -> Optional[str]:
        """Create a vendor directory Excel template."""
        if not self._check_openpyxl():
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors"

        headers = [
            "ID", "Name", "Contact", "Phone", "Email",
            "Address", "Specialty", "Payment Terms",
            "Lead Time", "Rating (1-5)", "Total Amount", "Notes"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            _style_header(cell)

        examples = [
            ["P001", "Example Stationery", "Sample Vendor A", "555-0104", "contact@examplestationery.example.com",
             "Commerce St 321", "Stationery and supplies", "Cash", "1-2 business days", 5, 0, "Fast delivery"],
            ["P002", "Example Cloud Services", "Sample Vendor B", "555-0201", "contact@examplecloud.example.com",
             "Digital Ave 100", "Cloud and hosting", "30 days", "Immediate", 4, 0, "Annual contract"],
            ["P003", "Example Transport", "Sample Vendor C", "555-0301", "contact@exampletransport.example.com",
             "Industrial Blvd 50", "Freight and courier", "15 days", "24-48 hours", 4, 0, ""],
        ]

        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                _style_body(cell)
                if col_idx == 11:
                    _format_currency(cell)

        widths = [8, 28, 20, 14, 28, 25, 28, 18, 20, 18, 14, 30]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:L100"

        wb.save(filepath)
        return filepath

    # Backward-compatible alias
    create_suppliers = create_vendors

    def create_cash_flow(self, filepath: str = "cash_flow.xlsx",
                         months: int = 6) -> Optional[str]:
        """Create a cash flow projection Excel template."""
        if not self._check_openpyxl():
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        ws["A1"] = "PROJECTED CASH FLOW"
        ws["A1"].font = Font(bold=True, size=18, color=COLOR_DARK)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        headers = ["Month", "Income", "Fixed Expenses", "Variable Expenses", "Balance", "Accumulated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            _style_header(cell)

        now = datetime.now()
        start_row = 4
        for i in range(months):
            row = start_row + i
            month = (now.month + i - 1) % 12 + 1
            year = now.year + (now.month + i - 1) // 12
            ws.cell(row=row, column=1, value=f"{year}-{month:02d}")
            ws.cell(row=row, column=2, value=0)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value=0)
            ws.cell(row=row, column=5, value=f"=B{row}-C{row}-D{row}")
            if i == 0:
                ws.cell(row=row, column=6, value=f"=E{row}")
            else:
                ws.cell(row=row, column=6, value=f"=F{row - 1}+E{row}")

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                _style_body(cell, align="center" if col == 1 else "right")
                if col in (2, 3, 4, 5, 6):
                    _format_currency(cell)

        widths = [14, 16, 16, 18, 16, 16]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[3].height = 25
        ws.freeze_panes = "A4"

        wb.save(filepath)
        return filepath


# Singleton
_templates_instance = None


def get_templates() -> Templates:
    """Get or create singleton Templates instance."""
    global _templates_instance
    if _templates_instance is None:
        _templates_instance = Templates()
    return _templates_instance
