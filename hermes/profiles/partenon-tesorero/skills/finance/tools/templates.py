"""
Partenon Tesorero — Excel Templates
Generate reusable Excel templates for budgets, providers and cash flow.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

# Optional dependency
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Partenon palette for Excel (dark blue, amber, cyan accents)
COLOR_DARK = "0B1F3A"
COLOR_ACCENT = "00D4FF"
COLOR_AMBER = "FFB800"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT = "E8F4F8"


def _style_header(cell):
    """Apply header style to a cell."""
    cell.font = Font(bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill(start_color=COLOR_DARK, end_color=COLOR_DARK, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _style_body(cell, align: str = "left"):
    """Apply body style to a cell."""
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


def _format_currency(cell):
    """Apply currency format."""
    cell.number_format = '$#,##0.00'


def _format_percent(cell):
    """Apply percentage format."""
    cell.number_format = '0.00%'


class Templates:
    """Generate finance Excel templates."""

    def __init__(self):
        self.errors: List[str] = []

    def _check_openpyxl(self) -> bool:
        if not OPENPYXL_AVAILABLE:
            self.errors.append("openpyxl no esta instalado")
            return False
        return True

    def crear_presupuesto(self, filepath: str = "presupuesto.xlsx",
                          periodo: str = None,
                          rubros: Optional[List[Dict[str, Any]]] = None) -> Optional[str]:
        """Create a budget vs actual Excel template."""
        if not self._check_openpyxl():
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto"

        periodo = periodo or datetime.now().strftime("%B %Y")

        # Title
        ws["A1"] = "PRESUPUESTO VS REAL"
        ws["A1"].font = Font(bold=True, size=18, color=COLOR_DARK)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws["A2"] = "Periodo:"
        ws["B2"] = periodo
        ws["A2"].font = Font(bold=True)

        # Headers
        headers = ["Rubro", "Presupuestado", "Real", "Diferencia", "% Variacion", "Observaciones"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            _style_header(cell)

        default_rubros = [
            {"rubro": "Materiales", "presupuestado": 30000},
            {"rubro": "Publicidad", "presupuestado": 12000},
            {"rubro": "Salarios", "presupuestado": 45000},
            {"rubro": "Renta", "presupuestado": 15000},
            {"rubro": "Servicios", "presupuestado": 5000},
            {"rubro": "Flete", "presupuestado": 5000},
        ]
        rubros = rubros or default_rubros

        start_row = 5
        for idx, item in enumerate(rubros):
            row = start_row + idx
            ws.cell(row=row, column=1, value=item.get("rubro", ""))
            ws.cell(row=row, column=2, value=item.get("presupuestado", 0))
            ws.cell(row=row, column=3, value=item.get("real", 0))
            ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
            ws.cell(row=row, column=5, value=f"=IF(B{row}=0,0,D{row}/B{row})")
            ws.cell(row=row, column=6, value=item.get("observaciones", ""))

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                _style_body(cell, align="left" if col in (1, 6) else "center")
                if col in (2, 3, 4):
                    _format_currency(cell)
                elif col == 5:
                    _format_percent(cell)

        # Totals
        total_row = start_row + len(rubros)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")

        for col in (2, 3, 4):
            cell = ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{total_row - 1})"
            )
            cell.font = Font(bold=True)
            _format_currency(cell)
            _style_body(cell, align="center")

        ws.cell(row=total_row, column=5, value=f"=IF(B{total_row}=0,0,D{total_row}/B{total_row})")
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        _format_percent(ws.cell(row=total_row, column=5))
        _style_body(ws.cell(row=total_row, column=5), align="center")

        # Widths
        widths = [28, 18, 14, 14, 15, 40]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[4].height = 25
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:F{total_row}"

        # Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Presupuestado vs Real"
        chart.y_axis.title = "Monto"
        chart.x_axis.title = "Rubro"

        data = Reference(ws, min_col=2, min_row=4, max_row=total_row - 1, max_col=3)
        cats = Reference(ws, min_col=1, min_row=5, max_row=total_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4

        ws.add_chart(chart, "H4")

        wb.save(filepath)
        return filepath

    def crear_proveedores(self, filepath: str = "proveedores.xlsx") -> Optional[str]:
        """Create a provider directory Excel template."""
        if not self._check_openpyxl():
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Proveedores"

        headers = [
            "ID", "Nombre", "Contacto", "Telefono", "Email",
            "Direccion", "Especialidad", "Condiciones de Pago",
            "Tiempo de Entrega", "Calificacion (1-5)", "Monto Total", "Notas"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            _style_header(cell)

        examples = [
            ["P001", "Papeleria Central", "Ana Garcia", "555-0104", "ana@papeleriacentral.com",
             "Calle Comercio 321", "Papeleria y consumibles", "Contado", "1-2 dias habiles", 5, 0, "Entrega rapida"],
            ["P002", "Servicios Cloud MX", "Carlos Ruiz", "555-0201", "carlos@cloudmx.com",
             "Av. Digital 100", "Cloud y hosting", "30 dias", "Inmediato", 4, 0, "Contrato anual"],
            ["P003", "Transportes del Norte", "Maria Lopez", "555-0301", "maria@tdnorte.com",
             "Blvd. Industrial 50", "Flete y mensajeria", "15 dias", "24-48 horas", 4, 0, ""],
        ]

        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                _style_body(cell)
                if col_idx == 11:
                    _format_currency(cell)

        widths = [8, 28, 20, 14, 28, 25, 28, 18, 20, 18, 14, 30]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:L100"

        wb.save(filepath)
        return filepath

    def crear_flujo_caja(self, filepath: str = "flujo_caja.xlsx",
                         meses: int = 6) -> Optional[str]:
        """Create a cash flow projection Excel template."""
        if not self._check_openpyxl():
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Flujo de Caja"

        ws["A1"] = "FLUJO DE CAJA PROYECTADO"
        ws["A1"].font = Font(bold=True, size=18, color=COLOR_DARK)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        headers = ["Mes", "Ingresos", "Gastos Fijos", "Gastos Variables", "Balance", "Acumulado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            _style_header(cell)

        now = datetime.now()
        start_row = 4
        for i in range(meses):
            row = start_row + i
            mes = (now.month + i - 1) % 12 + 1
            anio = now.year + (now.month + i - 1) // 12
            ws.cell(row=row, column=1, value=f"{anio}-{mes:02d}")
            ws.cell(row=row, column=2, value=0)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value=0)
            ws.cell(row=row, column=5, value=f"=B{row}-C{row}-D{row}")
            if i == 0:
                ws.cell(row=row, column=6, value=f"=E{row}")
            else:
                ws.cell(row=row, column=6, value=f"=F{row - 1}+E{row}")

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                _style_body(cell, align="center" if col == 1 else "right")
                if col in (2, 3, 4, 5, 6):
                    _format_currency(cell)

        widths = [14, 16, 16, 18, 16, 16]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[3].height = 25
        ws.freeze_panes = "A4"

        wb.save(filepath)
        return filepath


# Singleton
_templates_instance = None


def get_templates() -> Templates:
    """Get or create singleton Templates instance."""
    global _templates_instance
    if _templates_instance is None:
        _templates_instance = Templates()
    return _templates_instance
