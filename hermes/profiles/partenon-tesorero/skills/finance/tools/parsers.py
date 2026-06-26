"""
Partenon Tesorero - Expense Parsers
Extract and normalize expenses from Excel and CSV files.
"""

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

# Optional dependencies
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


CATEGORY_KEYWORDS = {
    "renta": ["renta", "arrendamiento", "oficina"],
    "nomina": ["nomina", "salario", "sueldo", "aguinaldo", "honorarios"],
    "tecnologia": ["cloud", "hosting", "dominio", "software", "suscripcion", "saas"],
    "marketing": ["publicidad", "ads", "campana", "facebook", "google ads", "impresion"],
    "materiales": ["material", "insumo", "mercancia", "producto", "inventario"],
    "logistica": ["flete", "envio", "mensajeria", "transporte", "gasolina"],
    "servicios": ["luz", "agua", "internet", "telefono", "limpieza", "seguridad"],
    "impuestos": ["iva", "isr", "impuesto", "tenencia", "licencia"],
    "mantenimiento": ["mantenimiento", "reparacion", "servicio tecnico"],
    "otros": [],
}

FIXED_KEYWORDS = [
    "renta", "arrendamiento", "nomina", "salario", "sueldo", "seguro",
    "internet", "telefono", "hosting", "dominio", "suscripcion", "saas",
    "luz", "agua", "gas", "limpieza", "vigilancia", "seguridad",
]


def normalize_text(value: Any) -> str:
    """Normalize text for matching."""
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.upper().split())


def normalize_amount(value: Any) -> Optional[float]:
    """Normalize a value to a float amount."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace("$", "").replace(",", "")
    text = text.replace("(", "-").replace(")", "")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_date(value: Any, default: Optional[datetime] = None) -> Optional[datetime]:
    """Normalize a value to a datetime."""
    if value is None:
        return default
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                pass
    return default


def infer_category(description: str) -> str:
    """Infer expense category from description."""
    text = normalize_text(description)
    for category, keywords in CATEGORY_KEYWORDS.items():
        if category == "otros":
            continue
        for keyword in keywords:
            if keyword in text:
                return category
    return "otros"


def infer_cost_type(description: str, category: Optional[str] = None) -> str:
    """Infer whether a cost is fixed or variable."""
    text = normalize_text(description)
    if category:
        text += " " + normalize_text(category)
    for keyword in FIXED_KEYWORDS:
        if keyword in text:
            return "fijo"
    return "variable"


def clean_value(value: Any) -> Any:
    """Clean a value for JSON output."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value


class ExpenseParser:
    """Parse expenses from Excel or CSV files."""

    COMMON_HEADERS = {
        "fecha": ["fecha", "date", "fecha de pago", "fecha operacion", "fecha transaccion"],
        "descripcion": ["descripcion", "concepto", "description", "movimiento", "detalle", "referencia"],
        "monto": ["monto", "importe", "cantidad", "amount", "total", "cargo", "abono"],
        "categoria": ["categoria", "category", "rubro", "tipo"],
        "proveedor": ["proveedor", "vendor", "supplier", "tienda", "comercio", " receptor"],
        "metodo": ["metodo", "forma de pago", "metodo de pago", "payment method"],
    }

    def __init__(self):
        self.errors: List[str] = []

    def parse_excel(self, filepath: str) -> List[Dict[str, Any]]:
        """Parse expenses from an Excel file."""
        if not OPENPYXL_AVAILABLE:
            self.errors.append("openpyxl no esta instalado")
            return []

        path = Path(filepath)
        if not path.exists():
            self.errors.append(f"Archivo no existe: {filepath}")
            return []

        expenses = []
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
                sheet_expenses = self._parse_rows(rows, source=f"{path.name}/{sheet_name}")
                expenses.extend(sheet_expenses)
            wb.close()
        except Exception as e:
            self.errors.append(f"Error leyendo Excel: {e}")

        return expenses

    def parse_csv(self, filepath: str, encoding: str = "utf-8") -> List[Dict[str, Any]]:
        """Parse expenses from a CSV file."""
        path = Path(filepath)
        if not path.exists():
            self.errors.append(f"Archivo no existe: {filepath}")
            return []

        expenses = []
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
            expenses = self._parse_rows(rows, source=path.name)
        except UnicodeDecodeError:
            return self.parse_csv(filepath, encoding="latin-1")
        except Exception as e:
            self.errors.append(f"Error leyendo CSV: {e}")

        return expenses

    def _parse_rows(self, rows: List[List[Any]], source: str) -> List[Dict[str, Any]]:
        """Parse rows into normalized expense dicts."""
        if not rows:
            return []

        header_row = rows[0]
        column_map = self._map_columns(header_row)

        if not column_map.get("monto"):
            # Try headerless format: look for columns with amounts
            column_map = self._infer_headerless_columns(rows)

        if not column_map.get("monto"):
            return []

        expenses = []
        for row in rows[1:]:
            if not any(row):
                continue

            expense = self._build_expense(row, column_map, source)
            if expense.get("monto") is not None and expense.get("monto") > 0:
                expenses.append(expense)

        return expenses

    def _map_columns(self, header_row: List[Any]) -> Dict[str, Optional[int]]:
        """Map standard fields to column indexes."""
        mapping: Dict[str, Optional[int]] = {key: None for key in self.COMMON_HEADERS}
        normalized_headers = [normalize_text(h) for h in header_row]

        for field, candidates in self.COMMON_HEADERS.items():
            for idx, header in enumerate(normalized_headers):
                for candidate in candidates:
                    if candidate.upper() in header:
                        mapping[field] = idx
                        break
                if mapping[field] is not None:
                    break

        return mapping

    def _infer_headerless_columns(self, rows: List[List[Any]]) -> Dict[str, Optional[int]]:
        """Infer columns when headers are missing."""
        mapping: Dict[str, Optional[int]] = {key: None for key in self.COMMON_HEADERS}
        max_cols = max(len(row) for row in rows) if rows else 0

        for col_idx in range(max_cols):
            values = [row[col_idx] if col_idx < len(row) else None for row in rows[1:]]
            numeric_count = sum(1 for v in values if normalize_amount(v) is not None)

            if numeric_count > len(values) * 0.5:
                mapping["monto"] = col_idx
                break

        # Assume first column is date, second is description if available
        if max_cols >= 2:
            mapping["descripcion"] = 1
            if max_cols >= 1:
                mapping["fecha"] = 0

        return mapping

    def _build_expense(self, row: List[Any], column_map: Dict[str, Optional[int]], source: str) -> Dict[str, Any]:
        """Build a normalized expense dict from a row."""
        def get(field: str) -> Any:
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        raw_description = get("descripcion") or ""
        description = str(raw_description).strip()
        if not description:
            description = "Sin descripcion"

        amount = normalize_amount(get("monto"))
        date_value = normalize_date(get("fecha"))
        category = get("categoria")
        provider = get("proveedor")
        method = get("metodo")

        if category is None or str(category).strip() == "":
            category = infer_category(description)

        cost_type = infer_cost_type(description, category)

        return {
            "fecha": clean_value(date_value),
            "descripcion": description,
            "monto": amount,
            "categoria": category,
            "proveedor": str(provider).strip() if provider else "",
            "metodo": str(method).strip() if method else "",
            "tipo": cost_type,
            "source": source,
            "raw": {str(k): clean_value(v) for k, v in zip(self.COMMON_HEADERS.keys(), [get(f) for f in self.COMMON_HEADERS])},
        }

    def detect_duplicates(self, expenses: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Detect likely duplicate expenses."""
        seen: Dict[str, List[Dict[str, Any]]] = {}
        duplicates = []

        for expense in expenses:
            key = f"{expense.get('fecha')}|{expense.get('descripcion')}|{expense.get('monto')}"
            if key in seen:
                duplicates.append(expense)
            seen.setdefault(key, []).append(expense)

        return duplicates

    def detect_anomalies(self, expenses: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Detect anomalies in expenses."""
        anomalies = []
        for expense in expenses:
            monto = expense.get("monto")
            if monto is None or monto <= 0:
                anomalies.append({
                    "tipo": "monto invalido",
                    "expense": expense,
                    "razon": "El monto es nulo, cero o negativo",
                })
            if not expense.get("descripcion") or expense.get("descripcion") == "Sin descripcion":
                anomalies.append({
                    "tipo": "descripcion vacia",
                    "expense": expense,
                    "razon": "Falta descripcion",
                })
        return anomalies


# Singleton
_parser_instance = None


def get_parser() -> ExpenseParser:
    """Get or create singleton ExpenseParser instance."""
    global _parser_instance
    if _parser_instance is None:
        _parser_instance = ExpenseParser()
    return _parser_instance
