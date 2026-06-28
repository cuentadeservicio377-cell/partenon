"""
Scribe Demo: generates a sample expense workbook and runs a simple Partenon finance audit.
Run: python3 scripts/demo_scribe.py
"""

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "templates" / "google-sheet-base"))

from openpyxl import Workbook
from finance_sheet import create_finance_sheet


def add_sample_expenses(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(path)

    # Fixed Expenses
    ws = wb["Fixed Expenses"]
    for row in [
        [datetime(2026, 6, 1), "Coworking office", 450.0, "Monthly", "WeWork"],
        [datetime(2026, 6, 1), "Accounting software", 39.0, "Monthly", "QuickBooks"],
        [datetime(2026, 6, 1), "AWS security", 120.0, "Monthly", "Amazon Web Services"],
    ]:
        ws.append(row)

    # Variable Expenses
    ws = wb["Variable Expenses"]
    for row in [
        [datetime(2026, 6, 3), "Meta advertising", 250.0, "Marketing", "Meta"],
        [datetime(2026, 6, 5), "Freelance design", 600.0, "Design", "Upwork"],
        [datetime(2026, 6, 8), "Client travel", 180.0, "Operations", "Uber"],
    ]:
        ws.append(row)

    # Income
    ws = wb["Income"]
    for row in [
        [datetime(2026, 6, 2), "Consulting retainer", 2500.0, "Acme Inc.", "Wire transfer"],
        [datetime(2026, 6, 10), "Web development", 1500.0, "Beta Labs", "Stripe"],
    ]:
        ws.append(row)

    # Suppliers
    ws = wb["Suppliers"]
    for row in [
        ["WeWork", "Coworking", "wework.com", 450.0, 4],
        ["QuickBooks", "Accounting", "quickbooks.intuit.com", 39.0, 5],
        ["Amazon Web Services", "Infrastructure", "aws.amazon.com", 120.0, 5],
        ["Meta", "Advertising", "business.meta.com", 250.0, 3],
        ["Upwork", "Talent", "upwork.com", 600.0, 4],
    ]:
        ws.append(row)

    wb.save(path)


def audit_workbook(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)

    def col_sum(sheet_name: str, col_index: int):
        ws = wb[sheet_name]
        total = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                val = row[col_index]
                if isinstance(val, (int, float)):
                    total += float(val)
            except Exception:
                pass
        return total

    income = col_sum("Income", 2)
    fixed = col_sum("Fixed Expenses", 2)
    variable = col_sum("Variable Expenses", 2)
    margin = income - fixed - variable

    report = {
        "timestamp": datetime.now(timezone.utc).isoformat() + "Z",
        "income": income,
        "fixed_expenses": fixed,
        "variable_expenses": variable,
        "margin": margin,
        "margin_pct": round(margin / income * 100, 2) if income else 0,
        "alerts": [],
    }

    if margin < 0:
        report["alerts"].append("Negative margin: expenses exceed income.")
    if fixed > income * 0.5:
        report["alerts"].append("Fixed expenses exceed 50% of income.")
    if variable > income * 0.3:
        report["alerts"].append("Variable expenses exceed 30% of income.")

    return report


def main():
    root = Path(__file__).resolve().parents[1]
    sample = root / "data" / "sample_expenses.xlsx"
    sample.parent.mkdir(parents=True, exist_ok=True)

    create_finance_sheet(sample)
    add_sample_expenses(sample)
    report = audit_workbook(sample)

    report_path = root / "data" / "sample_expenses_report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))

    print("=== Partenon Scribe Demo ===")
    print(f"Workbook: {sample}")
    print(f"Report: {report_path}")
    print(json.dumps(report, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
