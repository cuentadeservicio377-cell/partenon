"""
Demo Tesorero: generates a sample expense workbook and runs a simple Partenon finance audit.
Run: python scripts/demo_tesorero.py
"""

import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "templates" / "google-sheet-base"))

from openpyxl import Workbook
from finance_sheet import create_finance_sheet


def add_sample_expenses(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(path)

    # Gastos Fijos
    ws = wb["Gastos Fijos"]
    for row in [
        [datetime(2026, 6, 1), "Oficina coworking", 450.0, "Mensual", "WeWork"],
        [datetime(2026, 6, 1), "Software contable", 39.0, "Mensual", "QuickBooks"],
        [datetime(2026, 6, 1), "Seguridad AWS", 120.0, "Mensual", "Amazon Web Services"],
    ]:
        ws.append(row)

    # Gastos Variables
    ws = wb["Gastos Variables"]
    for row in [
        [datetime(2026, 6, 3), "Publicidad Meta", 250.0, "Marketing", "Meta"],
        [datetime(2026, 6, 5), "Freelance diseno", 600.0, "Diseno", "Upwork"],
        [datetime(2026, 6, 8), "Viaje cliente", 180.0, "Operaciones", "Uber"],
    ]:
        ws.append(row)

    # Ingresos
    ws = wb["Ingresos"]
    for row in [
        [datetime(2026, 6, 2), "Retainer consultoria", 2500.0, "Acme S.A.", "Transferencia"],
        [datetime(2026, 6, 10), "Desarrollo web", 1500.0, "Beta Labs", "Stripe"],
    ]:
        ws.append(row)

    # Proveedores
    ws = wb["Proveedores"]
    for row in [
        ["WeWork", "Coworking", "wework.com", 450.0, 4],
        ["QuickBooks", "Contabilidad", "quickbooks.intuit.com", 39.0, 5],
        ["Amazon Web Services", "Infraestructura", "aws.amazon.com", 120.0, 5],
        ["Meta", "Publicidad", "business.meta.com", 250.0, 3],
        ["Upwork", "Talento", "upwork.com", 600.0, 4],
    ]:
        ws.append(row)

    wb.save(path)


def audit_workbook(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)

    def col_sum(sheet_name: str, col_index: int):
        ws = wb[sheet_name]
        total = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                val = row[col_index]
                if isinstance(val, (int, float)):
                    total += float(val)
            except Exception:
                pass
        return total

    income = col_sum("Ingresos", 2)
    fixed = col_sum("Gastos Fijos", 2)
    variable = col_sum("Gastos Variables", 2)
    margin = income - fixed - variable

    report = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "ingresos": income,
        "gastos_fijos": fixed,
        "gastos_variables": variable,
        "margen": margin,
        "margen_pct": round(margin / income * 100, 2) if income else 0,
        "alertas": [],
    }

    if margin < 0:
        report["alertas"].append("Margen negativo: gastos superan ingresos.")
    if fixed > income * 0.5:
        report["alertas"].append("Gastos fijos superan el 50% de los ingresos.")
    if variable > income * 0.3:
        report["alertas"].append("Gastos variables superan el 30% de los ingresos.")

    return report


def main():
    root = Path(__file__).resolve().parents[1]
    sample = root / "data" / "sample_gastos.xlsx"
    sample.parent.mkdir(parents=True, exist_ok=True)

    create_finance_sheet(sample)
    add_sample_expenses(sample)
    report = audit_workbook(sample)

    report_path = root / "data" / "sample_gastos_report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))

    print("=== Partenon Tesorero Demo ===")
    print(f"Workbook: {sample}")
    print(f"Reporte: {report_path}")
    print(json.dumps(report, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
