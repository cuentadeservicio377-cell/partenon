"""
Generates a Partenon finance master spreadsheet template.
Can be used locally with openpyxl or uploaded to Google Sheets.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ACCENT = "00D4FF"
BG = "050505"
TEXT = "E8E8ED"
MUTED = "6B6B78"


def style_header(cell):
    cell.font = Font(name="Space Grotesk", size=11, bold=True, color=BG)
    cell.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_cell(cell, bold=False):
    cell.font = Font(name="Geist", size=10, color=TEXT, bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def create_finance_sheet(output_path: Path):
    wb = Workbook()

    # Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    headers = ["Indicador", "Valor", "Nota"]
    ws.append(headers)
    for cell in ws[1]:
        style_header(cell)

    rows = [
        ["Ingresos mes actual", "=SUM(Ingresos[C])", "Suma de hoja Ingresos"],
        ["Gastos fijos mes", "=SUMIF('Gastos Fijos'[C],\"<>\")", "Costos recurrentes"],
        ["Gastos variables mes", "=SUMIF('Gastos Variables'[C],\"<>\")", "Costos proporcionales"],
        ["Margen estimado", "=B2-B3-B4", "Ingresos - gastos"],
        ["Proveedores activos", "=COUNTA(Proveedores[A])-1", ""],
    ]
    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            style_cell(cell)

    # Ingresos
    ws = wb.create_sheet("Ingresos")
    ws.append(["Fecha", "Concepto", "Monto", "Cliente", "Canal"])
    for cell in ws[1]:
        style_header(cell)

    # Gastos Fijos
    ws = wb.create_sheet("Gastos Fijos")
    ws.append(["Fecha", "Concepto", "Monto", "Frecuencia", "Proveedor"])
    for cell in ws[1]:
        style_header(cell)

    # Gastos Variables
    ws = wb.create_sheet("Gastos Variables")
    ws.append(["Fecha", "Concepto", "Monto", "Categoria", "Proveedor"])
    for cell in ws[1]:
        style_header(cell)

    # Proveedores
    ws = wb.create_sheet("Proveedores")
    ws.append(["Nombre", "Servicio", "Contacto", "Pago habitual", "Calificacion"])
    for cell in ws[1]:
        style_header(cell)

    for sheet in wb.worksheets:
        auto_width(sheet)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    output = Path(__file__).resolve().parents[2] / "data" / "partenon-finance-template.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    create_finance_sheet(output)
    print(f"Created {output}")
